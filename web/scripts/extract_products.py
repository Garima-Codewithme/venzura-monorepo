import json
import re
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
INPUT_FILE = ROOT / "data-source" / "product-list.xlsx"
OUTPUT_DIR = ROOT / "src" / "data" / "generated"
OUTPUT_FILE = OUTPUT_DIR / "products.json"


def clean_text(value):
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def slugify(value):
    value = clean_text(value).lower()
    value = value.replace("&", " and ")
    value = re.sub(r"[^a-z0-9]+", "-", value)
    value = re.sub(r"-+", "-", value).strip("-")
    return value


def detect_heading(text):
    t = clean_text(text).lower()

    if "dry injection" in t:
        return "Dry Injection"
    if "liquid injection" in t:
        return "Liquid Injection"
    if "pre-filled syringe" in t or "pre-filled" in t or "pre filled" in t:
        return "Pre-Filled Syringe"
    if t == "p.p.i" or "p.p.i" in t:
        return "PPI"
    if "infusion iv" in t:
        return "Infusion IVs"

    return None


def infer_category(name, packaging, current_category):
    name_l = clean_text(name).lower()
    packaging_l = clean_text(packaging).lower()

    if "lyophilized" in packaging_l or "lyophilized" in name_l or "lyopholized" in packaging_l:
        return "Lyophilized (Cake Form)"

    if "prefilled syringe" in packaging_l or "pre filled syringe" in packaging_l:
        return "Pre-Filled Syringe"

    if "infusion" in name_l or "infusion" in packaging_l:
        return "Infusion IVs"

    return current_category


def map_dosage_form(category):
    mapping = {
        "Dry Injection": "Dry Injection",
        "Liquid Injection": "Liquid Injection",
        "Lyophilized (Cake Form)": "Lyophilized Injection",
        "Pre-Filled Syringe": "Pre-Filled Syringe",
        "PPI": "Injection",
        "Infusion IVs": "Infusion",
    }
    return mapping.get(category, "Pharma Product")


def is_noise(text):
    t = clean_text(text).lower()

    if not t:
        return True

    bad_parts = [
        "venzura medcor",
        "gst no",
        "address",
        "mob no",
        "haryana",
        "ambala",
    ]

    return any(part in t for part in bad_parts)


def main():
    if not INPUT_FILE.exists():
        print(f"File not found: {INPUT_FILE}")
        return

    wb = load_workbook(INPUT_FILE, data_only=True)
    ws = wb["Sheet1"]

    products = []
    slug_set = set()
    current_category = None

    for row_idx in range(1, ws.max_row + 1):
        col_a = ws.cell(row=row_idx, column=1).value
        col_b = ws.cell(row=row_idx, column=2).value
        col_c = ws.cell(row=row_idx, column=3).value

        text_a = clean_text(col_a)
        text_b = clean_text(col_b)
        text_c = clean_text(col_c)

        heading = detect_heading(text_a) or detect_heading(text_b) or detect_heading(text_c)
        if heading:
            current_category = heading
            continue

        product_name = text_b
        packaging = text_c

        if not clean_text(product_name):
            continue
        if is_noise(product_name):
            continue
        if detect_heading(product_name):
            continue

        category = infer_category(product_name, packaging, current_category)

        if not category:
            continue

        base_slug = slugify(f"{category}-{product_name}")
        if not base_slug:
            continue

        final_slug = base_slug
        counter = 2
        while final_slug in slug_set:
            final_slug = f"{base_slug}-{counter}"
            counter += 1

        slug_set.add(final_slug)

        products.append(
            {
                "name": clean_text(product_name),
                "slug": final_slug,
                "division": "Pharma",
                "category": category,
                "dosageForm": map_dosage_form(category),
                "shortDescription": f"{clean_text(product_name)} from Venzura Medcor product portfolio.",
                "composition": "",
                "packaging": clean_text(packaging),
                "featured": False,
                "sourceSheet": "Sheet1",
                "sourceRow": row_idx,
            }
        )

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    OUTPUT_FILE.write_text(json.dumps(products, indent=2), encoding="utf-8")

    summary = {}
    for item in products:
        summary[item["category"]] = summary.get(item["category"], 0) + 1

    print(f"Saved {len(products)} products to:")
    print(OUTPUT_FILE)
    print("\nCategory summary:")
    for key, value in summary.items():
        print(f"- {key}: {value}")


if __name__ == "__main__":
    main()