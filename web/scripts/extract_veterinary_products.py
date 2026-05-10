import json
import re
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
INPUT_FILE_A = ROOT / "data-source" / "veterinary-a.xlsx"
INPUT_FILE_B = ROOT / "data-source" / "veterinary-b.xlsx"
OUTPUT_DIR = ROOT / "src" / "data" / "generated"
OUTPUT_FILE = OUTPUT_DIR / "veterinary-products.json"

CATEGORY_ID = "category_veterinary"
CATEGORY_SLUG = "veterinary"
DIVISION = "Veterinary"

SUBCATEGORY_IDS = {
    "liquid_tonics": "subcategory_veterinary_liquid_tonics",
    "bolus": "subcategory_veterinary_bolus_tablets_kits",
    "ointment": "subcategory_veterinary_ointment",
    "spray": "subcategory_veterinary_spray",
    "shampoo": "subcategory_veterinary_shampoo",
    "injections": "subcategory_veterinary_injections",
    "feed_supplements": "subcategory_veterinary_feed_supplements",
}

def clean_text(value):
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()

def slugify(value):
    value = clean_text(value).lower()
    value = value.replace("&", " and ")
    value = re.sub(r"[^a-z0-9]+", "-", value)
    value = re.sub(r"-+", "-", value).strip("-")
    return value

def title_case_from_label(label):
    words = clean_text(label).lower().split()
    return " ".join(word.capitalize() for word in words)

def derive_name_from_composition(composition, packaging=""):
    text = clean_text(composition)

    # Prefer bracket label at end like (LIVER TONIC), (CALCIUM), etc.
    matches = re.findall(r"\(([^()]*)\)", text)
    if matches:
        candidate = clean_text(matches[-1])
        if 2 <= len(candidate) <= 60:
            return title_case_from_label(candidate)

    # Remove common prefixes
    text = re.sub(r"^each\s*\d*\s*ml\s*contains:\s*", "", text, flags=re.I)
    text = re.sub(r"^each\s*\d*\s*gms\s*contains:\s*", "", text, flags=re.I)
    text = re.sub(r"^each\s*\d*\s*gm\s*contains:\s*", "", text, flags=re.I)

    # Take first 8 words
    words = text.split()
    short = " ".join(words[:8]).strip(" ,.-")
    short = title_case_from_label(short)

    if not short:
        short = "Veterinary Product"

    return short

def make_short_description(name, subcategory_title, packaging):
    packaging_text = f" Available in {packaging}." if packaging else ""
    return f"{name} under {subcategory_title} from Venzura Medcor veterinary portfolio.{packaging_text}"

def guess_dosage_form(subcategory_title, packaging):
    sub = clean_text(subcategory_title).lower()
    pack = clean_text(packaging).lower()

    if "injection" in sub:
        return "Injection"
    if "bolus" in sub or "tablet" in sub:
        return "Bolus / Tablet"
    if "ointment" in sub:
        return "Ointment"
    if "spray" in sub:
        return "Spray"
    if "shampoo" in sub:
        return "Shampoo"
    if "liquid" in sub:
        return "Liquid"
    if "feed" in sub:
        return "Feed Supplement"

    if "vial" in pack:
        return "Injection"
    if "bolus" in pack:
        return "Bolus"
    if "bottle" in pack or "ltr" in pack or "ml" in pack:
        return "Liquid"

    return "Veterinary Product"

def normalize_key(composition, packaging, subcategory_id):
    c = re.sub(r"[^a-z0-9]+", "", clean_text(composition).lower())
    p = re.sub(r"[^a-z0-9]+", "", clean_text(packaging).lower())
    s = clean_text(subcategory_id).lower()
    return f"{s}|{c}|{p}"

def extract_from_file_a():
    # File A:
    # col3 = composition/description
    # col4 = packaging
    # row 23 = BOLUS SECTION
    products = []

    wb = load_workbook(INPUT_FILE_A, data_only=True)
    ws = wb["Sheet1"]

    current_subcategory_id = SUBCATEGORY_IDS["liquid_tonics"]
    current_subcategory_title = "Liquid Tonics & Supplements"

    for row_idx in range(2, ws.max_row + 1):
        composition = clean_text(ws.cell(row=row_idx, column=3).value)
        packaging = clean_text(ws.cell(row=row_idx, column=4).value)

        if not composition:
            continue

        if "bolus section" in composition.lower():
            current_subcategory_id = SUBCATEGORY_IDS["bolus"]
            current_subcategory_title = "Bolus / Tablets / Kits"
            continue

        name = derive_name_from_composition(composition, packaging)
        slug = slugify(f"{current_subcategory_title}-{name}-{packaging}")

        products.append({
            "name": name,
            "slug": slug,
            "categoryId": CATEGORY_ID,
            "categorySlug": CATEGORY_SLUG,
            "subcategoryId": current_subcategory_id,
            "subcategoryTitle": current_subcategory_title,
            "shortDescription": make_short_description(name, current_subcategory_title, packaging),
            "composition": composition,
            "dosageForm": guess_dosage_form(current_subcategory_title, packaging),
            "packaging": packaging,
            "division": DIVISION,
            "featured": False,
            "sourceFile": "veterinary-a.xlsx",
        })

    return products

def extract_from_file_b():
    # File B:
    # col2 = composition
    # col3/4 may contain packing
    products = []

    wb = load_workbook(INPUT_FILE_B, data_only=True)
    ws = wb["Sheet1"]

    current_subcategory_id = None
    current_subcategory_title = None

    heading_map = {
        "bolus / tablets / kits": ("Bolus / Tablets / Kits", SUBCATEGORY_IDS["bolus"]),
        "liquid section": ("Liquid Tonics & Supplements", SUBCATEGORY_IDS["liquid_tonics"]),
        "ointment": ("Ointment", SUBCATEGORY_IDS["ointment"]),
        "spray": ("Spray", SUBCATEGORY_IDS["spray"]),
        "shampoo": ("Shampoo", SUBCATEGORY_IDS["shampoo"]),
        "injections": ("Injections", SUBCATEGORY_IDS["injections"]),
        "feed supplements": ("Feed Supplements", SUBCATEGORY_IDS["feed_supplements"]),
    }

    for row_idx in range(2, ws.max_row + 1):
        col2 = clean_text(ws.cell(row=row_idx, column=2).value)
        col3 = clean_text(ws.cell(row=row_idx, column=3).value)
        col4 = clean_text(ws.cell(row=row_idx, column=4).value)
        col5 = clean_text(ws.cell(row=row_idx, column=5).value)

        if not col2 and not col3 and not col4 and not col5:
            continue

        normalized_heading = col2.lower()
        if normalized_heading in heading_map:
            current_subcategory_title, current_subcategory_id = heading_map[normalized_heading]
            continue

        if not current_subcategory_id:
            continue

        composition = col2
        packaging = col3 or col4 or col5

        if not composition:
            continue

        # Skip fake brand column-like text
        if composition.lower() in ["your brand", "your  brand"]:
            continue

        name = derive_name_from_composition(composition, packaging)
        slug = slugify(f"{current_subcategory_title}-{name}-{packaging}")

        products.append({
            "name": name,
            "slug": slug,
            "categoryId": CATEGORY_ID,
            "categorySlug": CATEGORY_SLUG,
            "subcategoryId": current_subcategory_id,
            "subcategoryTitle": current_subcategory_title,
            "shortDescription": make_short_description(name, current_subcategory_title, packaging),
            "composition": composition,
            "dosageForm": guess_dosage_form(current_subcategory_title, packaging),
            "packaging": packaging,
            "division": DIVISION,
            "featured": False,
            "sourceFile": "veterinary-b.xlsx",
        })

    return products

def dedupe_products(items):
    deduped = []
    seen = set()
    slug_counts = {}

    for item in items:
        key = normalize_key(item["composition"], item["packaging"], item["subcategoryId"])
        if key in seen:
            continue
        seen.add(key)

        base_slug = item["slug"] or slugify(item["name"])
        final_slug = base_slug

        if final_slug in slug_counts:
            slug_counts[final_slug] += 1
            final_slug = f"{final_slug}-{slug_counts[final_slug]}"
        else:
            slug_counts[final_slug] = 1

        item["slug"] = final_slug
        deduped.append(item)

    return deduped

def main():
    if not INPUT_FILE_A.exists():
        print(f"Missing file: {INPUT_FILE_A}")
        return

    if not INPUT_FILE_B.exists():
        print(f"Missing file: {INPUT_FILE_B}")
        return

    products_a = extract_from_file_a()
    products_b = extract_from_file_b()

    merged = products_a + products_b
    deduped = dedupe_products(merged)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    OUTPUT_FILE.write_text(json.dumps(deduped, indent=2), encoding="utf-8")

    summary = {}
    for item in deduped:
        summary[item["subcategoryTitle"]] = summary.get(item["subcategoryTitle"], 0) + 1

    print(f"File A products: {len(products_a)}")
    print(f"File B products: {len(products_b)}")
    print(f"After dedupe: {len(deduped)}")
    print(f"Saved merged veterinary products to:")
    print(OUTPUT_FILE)

    print("\nSubcategory summary:")
    for key, value in sorted(summary.items()):
        print(f"- {key}: {value}")

if __name__ == "__main__":
    main()