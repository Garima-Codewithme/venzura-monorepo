import json
import re
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
INPUT_FILE = ROOT / "data-source" / "cosmetic-range.xlsx"
OUTPUT_DIR = ROOT / "src" / "data" / "generated"
OUTPUT_FILE = OUTPUT_DIR / "cosmetic-products.json"

CATEGORY_ID = "category_cosmetic"
CATEGORY_SLUG = "cosmetic"
DIVISION = "Cosmetic"

SHEET_MAP = {
    "lotion": ("Lotions", "subcategory_cosmetic_lotions"),
    "face wash": ("Face Wash", "subcategory_cosmetic_face_wash"),
    "cream": ("Cream", "subcategory_cosmetic_cream"),
    "shampoo & conditioner": ("Shampoo & Conditioner", "subcategory_cosmetic_shampoo_conditioner"),
    "hair oil": ("Hair Oil", "subcategory_cosmetic_hair_oil"),
    "soap": ("Soap", "subcategory_cosmetic_soap"),
    "body wash": ("Body Wash", "subcategory_cosmetic_body_wash"),
    "face toner": ("Face Toner", "subcategory_cosmetic_face_toner"),
    "face serum": ("Face Serum", "subcategory_cosmetic_face_serum"),
    "body scub": ("Body Scrub", "subcategory_cosmetic_body_scrub"),
    "hair serum": ("Hair Serum", "subcategory_cosmetic_hair_serum"),
    "facial kit": ("Facial Kit", "subcategory_cosmetic_facial_kit"),
    "hair mask": ("Hair Mask", "subcategory_cosmetic_hair_mask"),
    "lip balm": ("Lip Balm", "subcategory_cosmetic_lip_balm"),
    "under eye cream": ("Under Eye Cream", "subcategory_cosmetic_under_eye_cream"),
    "Hair removal cream": ("Hair Removal Cream", "subcategory_cosmetic_hair_removal_cream"),
    "face mask": ("Face Mask", "subcategory_cosmetic_face_mask"),
    "face scrub": ("Face Scrub", "subcategory_cosmetic_face_scrub"),
    "skin care gel": ("Skin Care Gel", "subcategory_cosmetic_skin_care_gel"),
    "foaming fash wash": ("Foaming Face Wash", "subcategory_cosmetic_foaming_face_wash"),
    "hand wash": ("Hand Wash", "subcategory_cosmetic_hand_wash"),
    "intimate hygiene wash": ("Intimate Hygiene Wash", "subcategory_cosmetic_intimate_hygiene_wash"),
    "essitinal oil": ("Essential Oil", "subcategory_cosmetic_essential_oil"),
    "intimate product": ("Intimate Products", "subcategory_cosmetic_intimate_products"),
    "beard range": ("Beard Range", "subcategory_cosmetic_beard_range"),
    "baby body wash": ("Baby Body Wash", "subcategory_cosmetic_baby_body_wash"),
    "baby soap": ("Baby Soap", "subcategory_cosmetic_baby_soap"),
    "baby lotion": ("Baby Lotion", "subcategory_cosmetic_baby_lotion"),
    "baby shampoo": ("Baby Shampoo", "subcategory_cosmetic_baby_shampoo"),
    "baby hair oil": ("Baby Hair Oil", "subcategory_cosmetic_baby_hair_oil"),
    "baby massage oil": ("Baby Massage Oil", "subcategory_cosmetic_baby_massage_oil"),
    "baby powder": ("Baby Powder", "subcategory_cosmetic_baby_powder"),
}

HEADING_MAP = {
    "lotions": ("Lotions", "subcategory_cosmetic_lotions"),
    "lotion": ("Lotions", "subcategory_cosmetic_lotions"),
    "face wash": ("Face Wash", "subcategory_cosmetic_face_wash"),
    "cream": ("Cream", "subcategory_cosmetic_cream"),
    "shampoo and conditioner": ("Shampoo & Conditioner", "subcategory_cosmetic_shampoo_conditioner"),
    "shampoo conditioner": ("Shampoo & Conditioner", "subcategory_cosmetic_shampoo_conditioner"),
    "hair oil": ("Hair Oil", "subcategory_cosmetic_hair_oil"),
    "soap": ("Soap", "subcategory_cosmetic_soap"),
    "body wash": ("Body Wash", "subcategory_cosmetic_body_wash"),
    "face toner": ("Face Toner", "subcategory_cosmetic_face_toner"),
    "face serum": ("Face Serum", "subcategory_cosmetic_face_serum"),
    "body scrub": ("Body Scrub", "subcategory_cosmetic_body_scrub"),
    "hair serum": ("Hair Serum", "subcategory_cosmetic_hair_serum"),
    "facial kit": ("Facial Kit", "subcategory_cosmetic_facial_kit"),
    "hair mask": ("Hair Mask", "subcategory_cosmetic_hair_mask"),
    "lip balm": ("Lip Balm", "subcategory_cosmetic_lip_balm"),
    "under eye cream": ("Under Eye Cream", "subcategory_cosmetic_under_eye_cream"),
    "hair removal cream": ("Hair Removal Cream", "subcategory_cosmetic_hair_removal_cream"),
    "face mask": ("Face Mask", "subcategory_cosmetic_face_mask"),
    "face scrub": ("Face Scrub", "subcategory_cosmetic_face_scrub"),
    "skin care gel": ("Skin Care Gel", "subcategory_cosmetic_skin_care_gel"),
    "foaming face wash": ("Foaming Face Wash", "subcategory_cosmetic_foaming_face_wash"),
    "foaming fash wash": ("Foaming Face Wash", "subcategory_cosmetic_foaming_face_wash"),
    "hand wash": ("Hand Wash", "subcategory_cosmetic_hand_wash"),
    "intimate hygiene wash": ("Intimate Hygiene Wash", "subcategory_cosmetic_intimate_hygiene_wash"),
    "essential oil": ("Essential Oil", "subcategory_cosmetic_essential_oil"),
    "essitinal oil": ("Essential Oil", "subcategory_cosmetic_essential_oil"),
    "intimate products": ("Intimate Products", "subcategory_cosmetic_intimate_products"),
    "intimate product": ("Intimate Products", "subcategory_cosmetic_intimate_products"),
    "beard range": ("Beard Range", "subcategory_cosmetic_beard_range"),
    "baby body wash": ("Baby Body Wash", "subcategory_cosmetic_baby_body_wash"),
    "baby soap": ("Baby Soap", "subcategory_cosmetic_baby_soap"),
    "baby lotion": ("Baby Lotion", "subcategory_cosmetic_baby_lotion"),
    "baby shampoo": ("Baby Shampoo", "subcategory_cosmetic_baby_shampoo"),
    "baby hair oil": ("Baby Hair Oil", "subcategory_cosmetic_baby_hair_oil"),
    "baby massage oil": ("Baby Massage Oil", "subcategory_cosmetic_baby_massage_oil"),
    "baby powder": ("Baby Powder", "subcategory_cosmetic_baby_powder"),
}

def clean_text(value):
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()

def slugify(value):
    value = clean_text(value).lower()
    value = value.replace("&", " and ")
    value = re.sub(r"[^a-z0-9]+", "-", value)
    value = re.sub(r"-+", "-", value).strip("-")
    return value

def normalize_heading(value):
    text = clean_text(value).lower()
    text = text.replace("&", " and ")
    text = re.sub(r"^\d+\.\s*", "", text)
    text = re.sub(r"[^a-z0-9\s]+", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text

def resolve_heading(value):
    heading = normalize_heading(value)
    return HEADING_MAP.get(heading)

def build_short_description(name, subcategory_title):
    return f"{name} under {subcategory_title} from Venzura Medcor cosmetic range."

def dedupe_key(name, subcategory_id):
    return f"{slugify(name)}|{subcategory_id}"

def main():
    if not INPUT_FILE.exists():
        print(f"File not found: {INPUT_FILE}")
        return

    workbook = load_workbook(INPUT_FILE, data_only=True)
    products = []
    seen = set()
    slug_counts = {}

    for sheet_name in workbook.sheetnames:
        if sheet_name not in SHEET_MAP:
            continue

        default_subcategory_title, default_subcategory_id = SHEET_MAP[sheet_name]
        current_subcategory_title = default_subcategory_title
        current_subcategory_id = default_subcategory_id

        worksheet = workbook[sheet_name]

        for row_idx in range(1, worksheet.max_row + 1):
            raw_name = clean_text(worksheet.cell(row=row_idx, column=1).value)

            if not raw_name:
                continue

            # Sheet title row skip
            if slugify(raw_name) == slugify(default_subcategory_title):
                continue

            # Numbering/keyword heading ke hisaab se subcategory switch
            heading_match = resolve_heading(raw_name)
            if heading_match:
                current_subcategory_title, current_subcategory_id = heading_match
                continue

            key = dedupe_key(raw_name, current_subcategory_id)
            if key in seen:
                continue
            seen.add(key)

            base_slug = slugify(raw_name)
            final_slug = base_slug

            if final_slug in slug_counts:
                slug_counts[final_slug] += 1
                final_slug = f"{final_slug}-{slug_counts[final_slug]}"
            else:
                slug_counts[final_slug] = 1

            products.append(
                {
                    "name": raw_name,
                    "slug": final_slug,
                    "categoryId": CATEGORY_ID,
                    "categorySlug": CATEGORY_SLUG,
                    "subcategoryId": current_subcategory_id,
                    "subcategoryTitle": current_subcategory_title,
                    "shortDescription": build_short_description(raw_name, current_subcategory_title),
                    "composition": "",
                    "dosageForm": "Cosmetic Product",
                    "packaging": "",
                    "division": DIVISION,
                    "featured": False,
                    "sourceSheet": sheet_name,
                }
            )

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    OUTPUT_FILE.write_text(json.dumps(products, indent=2), encoding="utf-8")

    summary = {}
    for item in products:
        summary[item["subcategoryTitle"]] = summary.get(item["subcategoryTitle"], 0) + 1

    print(f"Saved {len(products)} cosmetic products to:")
    print(OUTPUT_FILE)

    print("\nSubcategory summary:")
    for key, value in sorted(summary.items()):
        print(f"- {key}: {value}")

if __name__ == "__main__":
    main()