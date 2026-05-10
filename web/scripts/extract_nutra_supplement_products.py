import json
import re
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
INPUT_FILE = ROOT / "data-source" / "nutra-supplement-products.xlsx"
OUTPUT_DIR = ROOT / "src" / "data" / "generated"
OUTPUT_FILE = OUTPUT_DIR / "nutra-supplement-products.json"

CATEGORY_ID = "category_nutra_supplement"
CATEGORY_SLUG = "nutra-supplement"
DIVISION = "Nutra Supplement"

def clean_text(value):
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()

def slugify(value):
    value = clean_text(value).lower()
    value = value.replace("&", " and ")
    value = re.sub(r"[^a-z0-9]+", "-", value)
    value = re.sub(r"-+", "-", value).strip("-")
    return value

def build_short_description(name, packaging):
    packaging_text = f" Available in {packaging}." if packaging else ""
    return f"{name} from Venzura Medcor nutra supplement portfolio.{packaging_text}"

def guess_dosage_form(packaging):
    text = clean_text(packaging).lower()

    if "tablet" in text:
        return "Tablet"
    if "capsule" in text:
        return "Capsule"
    if "syrup" in text:
        return "Syrup"
    if "sachet" in text:
        return "Sachet"
    if "powder" in text:
        return "Powder"
    if "softgel" in text:
        return "Softgel"
    if "drops" in text:
        return "Drops"
    if "liquid" in text:
        return "Liquid"
    return "Nutra Supplement"

def main():
    if not INPUT_FILE.exists():
        print(f"File not found: {INPUT_FILE}")
        return

    workbook = load_workbook(INPUT_FILE, data_only=True)
    worksheet = workbook["Sheet1"]

    generic_rows = []
    for row_idx in range(1, worksheet.max_row + 1):
        value = worksheet.cell(row=row_idx, column=1).value
        text = clean_text(value)
        if re.search(r"generic name", text, re.IGNORECASE):
            generic_rows.append(row_idx)

    products = []
    slug_set = set()

    for index, start_row in enumerate(generic_rows):
        end_row = generic_rows[index + 1] if index + 1 < len(generic_rows) else worksheet.max_row + 1

        name = clean_text(worksheet.cell(row=start_row + 1, column=1).value)

        composition_lines = []
        packaging = ""
        mode = None

        for row_idx in range(start_row + 2, end_row):
            text = clean_text(worksheet.cell(row=row_idx, column=1).value)

            if not text:
                continue

            lower_text = text.lower()

            if lower_text == "composition":
                mode = "composition"
                continue

            if lower_text == "packaging":
                mode = "packaging"
                continue

            if "generic name" in lower_text:
                break

            if mode == "composition":
                composition_lines.append(text.rstrip(","))
            elif mode == "packaging" and not packaging:
                packaging = text

        if not name:
            continue

        base_slug = slugify(name)
        if not base_slug:
            continue

        final_slug = base_slug
        counter = 2

        while final_slug in slug_set:
            final_slug = f"{base_slug}-{counter}"
            counter += 1

        slug_set.add(final_slug)

        composition = ", ".join(composition_lines)

        products.append(
            {
                "name": name,
                "slug": final_slug,
                "categoryId": CATEGORY_ID,
                "categorySlug": CATEGORY_SLUG,
                "subcategoryId": "",
                "shortDescription": build_short_description(name, packaging),
                "composition": composition,
                "dosageForm": guess_dosage_form(packaging),
                "packaging": packaging,
                "division": DIVISION,
                "featured": False,
            }
        )

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    OUTPUT_FILE.write_text(json.dumps(products, indent=2), encoding="utf-8")

    print(f"Saved {len(products)} nutra supplement products to:")
    print(OUTPUT_FILE)

if __name__ == "__main__":
    main()